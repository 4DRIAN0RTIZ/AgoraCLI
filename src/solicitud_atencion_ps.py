"""
solitud_atencion_ps.py
Copyleft (c) 2023 Oscar Adrian Ortiz Bustos

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation version 3 of the License.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

"""

import openpyxl
import sys
import os
import time  # Importar el módulo time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl.utils import range_boundaries
from utils import Clear, Colors
from datetime import datetime

class AtencionPsicologicaSolicitud:
    def __init__(self, driver, download_path):
        self.driver = driver
        self.download_path = download_path

    def solicitar_atencion_psicologica(self):
        Clear.pantalla()
        print("Solicitando atención psicológica...")
        self.driver.get("https://agora.utj.edu.mx/AtencionPsicologica/solicitud")
        try:
            element = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, '[onclick="DescargarFormato()"]'))
            )
            element.click()
            self.esperar_descarga()
            self.procesar_solicitud()
        except Exception as e:
            print(f"Error: {str(e)}")
            sys.exit(1)

    def esperar_descarga(self):
        # Elimina archivos anteriores
        files = os.listdir(self.download_path)
        for file in files:
            if file.startswith("Formato") and file.endswith(".xlsx"):
                os.remove(os.path.join(self.download_path, file))

        # Esperar hasta que un nuevo archivo se descargue
        while True:
            files = os.listdir(self.download_path)
            horario_files = [f for f in files if f.startswith("Formato") and f.endswith(".xlsx")]
            if horario_files:
                time.sleep(1)  # Asegurar que la descarga esté completamente finalizada
                break


    def open_file(self):
        files = os.listdir(self.download_path)
        horario_files = [f for f in files if f.startswith("Formato") and f.endswith(".xlsx")]

        if not horario_files:
            print("Error: No se encontró ningún archivo de horario.")
            sys.exit(1)

        # Seleccionar el archivo más reciente
        latest_file = max(horario_files, key=lambda f: os.path.getmtime(os.path.join(self.download_path, f)))

        # Confirmar que solo se procese el archivo más reciente
        for file in horario_files:
            if file != latest_file:
                os.remove(os.path.join(self.download_path, file))

        return os.path.join(self.download_path, latest_file)

    def procesar_solicitud(self):
        file_path = self.open_file()
        if file_path is None:
            print("Error: No se pudo encontrar el archivo del horario.")
            sys.exit(1)
        
        Clear.linea_anterior()
        
        # Cargar el archivo Excel con openpyxl
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Verificar si las celdas están combinadas y escribir valores
        def actualizar_celda(rango, valor):
            if any(rango in str(merged) for merged in ws.merged_cells.ranges):
                # Si está combinada, escribir en la celda superior izquierda
                min_col, min_row, max_col, max_row = range_boundaries(rango)
                ws.cell(row=min_row, column=min_col, value=valor)
            else:
                ws[rango].value = valor

        # Leer y asignar valores
        campoNombreCompleto = input("Nombre completo: ")
        campoMatricula = input("Matrícula: ")
        campoPrograma = input("Programa (TSU(1)/ING(2)): ")
        if campoPrograma == "1":
            posicionCeldaP = "I12"
            campoPrograma = "TSU"
        elif campoPrograma == "2":
            posicionCeldaP = "O12"
            campoPrograma = "ING"
        else:
            print("Error: Carrera no válida.")
            sys.exit(1)
        campoTurno = input("Turno: ")
        campoCarrera = input("Carrera: ")
        campoGrado = input("Grado: ")
        campoUnidadAcademica = input("Unidad Académica (Miravalle(1)/CCD(2)): ")
        if campoUnidadAcademica == "1":
            posicionCeldaU = "I17"
            campoUnidadAcademica = "Miravalle"
        elif campoUnidadAcademica == "2":
            posicionCeldaU = "I18"
            campoUnidadAcademica = "CCD"
        else:
            print("Error: Unidad Académica no válida.")
            sys.exit(1)
        campoTipo = input("Tipo de atención (Asesoría(1)/Canalización(2)): ")
        if campoTipo == "1":
            posicionCeldaT = "O17"
            campoTipo = "Asesoría"
        elif campoTipo == "2":
            posicionCeldaT = "O18"
            campoTipo = "Canalización"
        else:
            print("Error: Tipo de atención no válida.")
            sys.exit(1)
        campoPrimeraVez = input("¿Es la primera ocasión que recibes atención psicológica por parte de la UTJ? (S/N): ")
        if campoPrimeraVez.upper() == "S":
            posicionCeldaPV = "U20"
            campoPrimeraVez = "Sí"
        elif campoPrimeraVez.upper() == "N":
            posicionCeldaPV = "X20"
            campoPrimeraVez = "No"
        else:
            print("Error: Opción no válida.")
            sys.exit(1)
        campoMotivo = input("Motivo por el cual requieres la atención: ")
        campoNombreTutor = input("Nombre completo del tutor: ")
        campoFechaSolicitud = datetime.now().strftime("%d/%m/%Y")

        actualizar_celda("E10", campoNombreCompleto)  # Celda para el nombre completo
        actualizar_celda("T10", campoMatricula)  # Celda para la matrícula
        actualizar_celda(posicionCeldaP, "X")
        actualizar_celda("T12", campoTurno)
        actualizar_celda("C14", campoCarrera)
        actualizar_celda("W14", campoGrado)
        actualizar_celda(posicionCeldaU, "X")
        actualizar_celda(posicionCeldaT, "X")
        actualizar_celda(posicionCeldaPV, "X")
        actualizar_celda("A23", campoMotivo)
        actualizar_celda("E27", campoNombreTutor)
        actualizar_celda("U27", campoFechaSolicitud)

        # Guardar el archivo con los cambios
        nuevo_path = file_path.replace("Formato", "Solicitud")
        wb.save(nuevo_path)
        input_file = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '[type="file"]'))
        )
        input_file.send_keys(nuevo_path)
        comentariosTxt = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.ID, 'txtComentarios'))
        )
        extra = input("Deseas agregar comentarios adicionales? (S/N): ")
        if extra.upper() == "S":
            comentarios = input("Comentarios: ")
            comentariosTxt.send_keys(comentarios)
        btn_enviar = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.ID, 'btnEnviar'))
        )
        print("========================================")
        print("RESUMEN DE LA SOLICITUD")
        print(f"{Colors.colorize('Nombre completo:', Colors.GREEN)} {campoNombreCompleto}")
        print(f"{Colors.colorize('Matrícula:', Colors.GREEN)} {campoMatricula}")
        print(f"{Colors.colorize('Programa:', Colors.GREEN)} {campoPrograma}")
        print(f"{Colors.colorize('Carrera:', Colors.GREEN)} {campoCarrera}")
        print(f"{Colors.colorize('Grado:', Colors.GREEN)} {campoGrado}")
        print(f"{Colors.colorize('Turno:', Colors.GREEN)} {campoTurno}")
        print(f"{Colors.colorize('Unidad Académica:', Colors.GREEN)} {campoUnidadAcademica}")
        print(f"{Colors.colorize('Tipo de atención:', Colors.GREEN)} {campoTipo}")
        print(f"{Colors.colorize('Primera vez:', Colors.GREEN)} {campoPrimeraVez}")
        print(f"{Colors.colorize('Motivo:', Colors.GREEN)} {campoMotivo}")
        print(f"{Colors.colorize('Nombre del tutor:', Colors.GREEN)} {campoNombreTutor}")
        print(f"{Colors.colorize('Fecha de solicitud:', Colors.GREEN)} {campoFechaSolicitud}")
        print("========================================")
        if btn_enviar.is_enabled():
            option = input("¿Deseas enviar la solicitud? (S/N): ")
            if option.upper() == "S":
                btn_enviar.click()
                print("Solicitud enviada con éxito.")
            else:
                print("Solicitud no enviada.")
        else:
            print("Error: No se pudo enviar la solicitud.")
